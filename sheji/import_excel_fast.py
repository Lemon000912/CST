"""
快速导入Excel数据到PostgreSQL - 使用COPY命令
"""

import asyncio
import uuid
import json
import csv
import tempfile
import os
from datetime import datetime
from core.database import engine
from sqlalchemy import text
import openpyxl


def clean_value(value):
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ['none', 'null', 'nan', '']:
            return ""
        return value.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ')
    return str(value)


def truncate_text(text, max_length=500):
    if not text:
        return ""
    text = str(text)
    if len(text) > max_length:
        return text[:max_length-3] + '...'
    return text


def extract_doi(filename):
    if not filename:
        return ""
    parts = filename.split('.')
    if len(parts) >= 2:
        potential_doi = '.'.join(parts[1:])
        if '10.' in potential_doi:
            return potential_doi[:100]
    return ""


async def import_excel_fast():
    print("Loading Excel file...")
    wb = openpyxl.load_workbook('合金与金属材料_二级分类.xlsx')
    ws = wb.active
    
    total_rows = ws.max_row - 1
    print(f"Total records: {total_rows}")
    
    headers = [cell.value for cell in ws[1]]
    field_mapping = {
        '一级分类': 'category',
        'file': 'filename',
        'DOI': 'doi',
        '第一作者': 'first_author',
        '通讯作者': 'corresponding_author',
        'SPL对称相': 'symmetry_phase',
        'DSC结构描述符': 'structure_descriptor',
        'PRO属性': 'properties',
        'APL应用': 'applications',
        'SMT合成工艺方法': 'synthesis_method',
        'CMT表征方法': 'characterization_method',
        'MAT材料': 'material_name',
        '质检': 'quality_control',
    }
    
    # 创建CSV文件
    csv_file = tempfile.NamedTemporaryFile(mode='w', newline='', encoding='utf-8', delete=False, suffix='.csv')
    csv_writer = csv.writer(csv_file, delimiter='|', quotechar='"', quoting=csv.QUOTE_MINIMAL)
    
    # 写入CSV头
    csv_writer.writerow([
        'id', 'filename', 'relative_path', 'file_size', 'title', 'authors', 'doi',
        'abstract', 'keywords', 'publish_year',
        'material_name', 'symmetry_phase', 'structure_descriptor',
        'properties', 'applications', 'synthesis_method',
        'characterization_method', 'quality_control',
        'first_author', 'corresponding_author',
        'category', 'is_processed', 'is_indexed', 'file_hash',
        'created_at', 'updated_at'
    ])
    
    now = datetime.now().isoformat()
    valid_count = 0
    
    print("Converting to CSV...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = {}
            for i, header in enumerate(headers):
                if header in field_mapping:
                    data[field_mapping[header]] = clean_value(row[i]) if i < len(row) else ""
            
            if not data.get('material_name'):
                continue
            
            if not data.get('doi') and data.get('filename'):
                data['doi'] = extract_doi(data['filename'])
            
            csv_writer.writerow([
                str(uuid.uuid4()),
                truncate_text(data.get('filename', ''), 255),
                f"alloy_metallic/{truncate_text(data.get('filename', ''), 255)}",
                0,
                truncate_text(data.get('material_name', ''), 500),
                json.dumps([data.get('first_author')] if data.get('first_author') else [], ensure_ascii=False),
                truncate_text(data.get('doi', ''), 100),
                "",
                json.dumps([], ensure_ascii=False),
                "",
                truncate_text(data.get('material_name', ''), 200),
                truncate_text(data.get('symmetry_phase', ''), 200),
                truncate_text(data.get('structure_descriptor', ''), 500),
                truncate_text(data.get('properties', ''), 1000),
                truncate_text(data.get('applications', ''), 500),
                truncate_text(data.get('synthesis_method', ''), 500),
                truncate_text(data.get('characterization_method', ''), 500),
                truncate_text(data.get('quality_control', ''), 200),
                truncate_text(data.get('first_author', ''), 100),
                truncate_text(data.get('corresponding_author', ''), 100),
                'alloy_metallic',
                'true',
                'true',
                str(uuid.uuid4()),
                now,
                now
            ])
            valid_count += 1
            
            if valid_count % 1000 == 0:
                print(f"  Processed: {valid_count}")
                
        except Exception as e:
            print(f"Error at row {row_idx}: {e}")
            continue
    
    csv_file.close()
    print(f"CSV created: {valid_count} valid records")
    
    # 使用COPY命令导入
    print("Importing to database using COPY...")
    async with engine.begin() as conn:
        # 先清空现有数据
        await conn.execute(text("DELETE FROM pdf_index WHERE category = 'alloy_metallic'"))
        print("Cleared existing data")
        
        # 使用COPY
        with open(csv_file.name, 'r', encoding='utf-8') as f:
            content = f.read()
            # 使用psycopg2的copy_expert
            import asyncpg
            raw_conn = await conn.get_raw_connection()
            asyncpg_conn = raw_conn.driver_connection
            
            await asyncpg_conn.copy_to_table(
                'pdf_index',
                source=csv_file.name,
                delimiter='|',
                format='csv',
                header=True
            )
        
        # 更新计数
        await conn.execute(
            text("""
                UPDATE material_categories 
                SET paper_count = (SELECT COUNT(*) FROM pdf_index WHERE category = 'alloy_metallic')
                WHERE code = 'alloy_metallic'
            """)
        )
    
    # 清理临时文件
    os.unlink(csv_file.name)
    
    print(f"\nImport completed! {valid_count} records imported")


if __name__ == "__main__":
    print("Starting fast Excel import...")
    print("=" * 60)
    asyncio.run(import_excel_fast())
