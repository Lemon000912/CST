"""
导入所有9大材料类别的数据
支持从Excel文件导入或生成模拟数据
"""

import asyncio
import uuid
import json
import os
from datetime import datetime
from core.database import engine
from sqlalchemy import text
import openpyxl


# 9大材料类别定义
MATERIAL_CATEGORIES = {
    'alloy_metallic': {
        'name': '合金/金属材料',
        'excel_file': '合金与金属材料_二级分类.xlsx',
        'sample_materials': ['Aluminum Alloy', 'Titanium Alloy', 'Steel', 'Copper Alloy', 'Nickel Alloy',
                           'Magnesium Alloy', 'Zinc Alloy', 'Brass', 'Bronze', 'Stainless Steel']
    },
    'amorphous_glass': {
        'name': '非晶玻璃',
        'excel_file': '非晶玻璃_二级分类.xlsx',
        'sample_materials': ['Metallic Glass', 'Chalcogenide Glass', 'Oxide Glass', 'Bulk Metallic Glass',
                           'Amorphous Silicon', 'Amorphous Carbon', 'Glass Ceramic', 'Optical Glass',
                           'Fiber Glass', 'Bioactive Glass']
    },
    'ceramic_structural': {
        'name': '结构陶瓷',
        'excel_file': '结构陶瓷_二级分类.xlsx',
        'sample_materials': ['Alumina Ceramic', 'Zirconia Ceramic', 'Silicon Carbide', 'Silicon Nitride',
                           'Boron Carbide', 'Tungsten Carbide', 'Titanium Carbide', 'Ceramic Matrix Composite',
                           'Piezoelectric Ceramic', 'Ferroelectric Ceramic']
    },
    'composite_multiphase': {
        'name': '多相复合材料',
        'excel_file': '多相复合材料_二级分类.xlsx',
        'sample_materials': ['Carbon Fiber Composite', 'Glass Fiber Composite', 'Metal Matrix Composite',
                           'Ceramic Matrix Composite', 'Polymer Matrix Composite', 'Nanocomposite',
                           'Hybrid Composite', 'Laminated Composite', 'Particulate Composite',
                           'Fiber Reinforced Polymer']
    },
    'nanomaterials_lowdim': {
        'name': '低维纳米材料',
        'excel_file': '低维纳米材料_二级分类.xlsx',
        'sample_materials': ['Graphene', 'Carbon Nanotube', 'Quantum Dot', 'Nanowire', 'Nanoparticle',
                           '2D Material', 'MXene', 'MoS2', 'Black Phosphorus', 'Nanorod',
                           'Nanoplate', 'Core-shell Nanoparticle']
    },
    'optical_optoelectronic': {
        'name': '光电材料',
        'excel_file': '光电材料_二级分类.xlsx',
        'sample_materials': ['GaAs', 'InP', 'GaN', 'ZnO', 'CdTe', 'Perovskite Solar Cell',
                           'OLED Material', 'Quantum Well', 'Photonic Crystal', 'Optical Fiber',
                           'LED Phosphor', 'Infrared Detector Material']
    },
    'polymer_soft_matter': {
        'name': '软物质高分子',
        'excel_file': '软物质高分子_二级分类.xlsx',
        'sample_materials': ['Polyethylene', 'Polypropylene', 'Polystyrene', 'PVC', 'PET',
                           'Nylon', 'Polycarbonate', 'Polyurethane', 'Silicone', 'Hydrogel',
                           'Liquid Crystal Polymer', 'Conductive Polymer', 'Biopolymer']
    },
    'solid_state_ionic': {
        'name': '固态离子材料',
        'excel_file': '固态离子材料_二级分类.xlsx',
        'sample_materials': ['Solid Electrolyte', 'Li-ion Conductor', 'Oxygen Ion Conductor',
                           'Proton Conductor', 'Sodium Ion Conductor', 'Solid Oxide Fuel Cell',
                           'All-solid-state Battery', 'Ionic Liquid', 'Polymer Electrolyte',
                           'Ceramic Electrolyte']
    },
    'surface_thin_film': {
        'name': '表面与薄膜材料',
        'excel_file': '表面与薄膜材料_二级分类.xlsx',
        'sample_materials': ['Thin Film Coating', 'DLC Coating', 'Anti-reflection Coating',
                           'Superhydrophobic Coating', 'Corrosion Resistant Coating',
                           'PVD Coating', 'CVD Coating', 'Atomic Layer Deposition',
                           'Self-assembled Monolayer', 'Nanostructured Surface']
    }
}

# 11个要素的Excel列名映射
ELEMENT_MAPPING = {
    'file': 'filename',
    'DOI': 'doi',
    '第一作者': 'first_author',
    '通讯作者': 'corresponding_author',
    'SPL对称相': 'symmetry_phase',
    'DSC结构描述符': 'structure_descriptor',
    'PRO属性': 'properties',
    'APL应用': 'applications',
    'SMT合成工艺方法': 'synthesis_method',
    'CMT表征方法': 'characterization_method',
    'MAT材料': 'material_name',
    '质检': 'quality_control',
}


def clean_value(value):
    """清理值"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ['none', 'null', 'nan', '', 'n/a']:
            return None
        return value
    return str(value) if value else None


def truncate_text(text, max_length=500):
    """截断文本"""
    if not text:
        return None
    text = str(text)
    if len(text) > max_length:
        return text[:max_length-3] + '...'
    return text


async def import_from_excel(category_code, excel_path):
    """从Excel文件导入数据"""
    print(f"\n{'='*60}")
    print(f"导入类别: {MATERIAL_CATEGORIES[category_code]['name']}")
    print(f"文件: {excel_path}")
    print('='*60)
    
    if not os.path.exists(excel_path):
        print(f"文件不存在: {excel_path}")
        return 0, 0
    
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        
        total = ws.max_row - 1
        print(f"总行数: {total}")
        
        headers = [cell.value for cell in ws[1]]
        
        # 获取原始连接
        async with engine.connect() as conn:
            raw_conn = await conn.get_raw_connection()
            asyncpg_conn = raw_conn.driver_connection
            
            # 清空该类别的现有数据
            await asyncpg_conn.execute(
                "DELETE FROM pdf_index WHERE category = $1",
                category_code
            )
            print(f"已清空 {category_code} 的现有数据")
            
            # 准备插入语句
            stmt = """
                INSERT INTO pdf_index (
                    id, filename, relative_path, file_size, title, authors, doi,
                    abstract, keywords, publish_year,
                    material_name, symmetry_phase, structure_descriptor,
                    properties, applications, synthesis_method,
                    characterization_method, quality_control,
                    first_author, corresponding_author,
                    category, is_processed, is_indexed, file_hash,
                    created_at, updated_at
                ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19, $20, $21, $22, $23, $24, NOW(), NOW())
            """
            
            imported = 0
            errors = 0
            
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    data = {ELEMENT_MAPPING.get(h): clean_value(row[i]) if i < len(row) else None 
                            for i, h in enumerate(headers) if h in ELEMENT_MAPPING}
                    
                    if not data.get('material_name'):
                        continue
                    
                    # 提取DOI
                    doi = data.get('doi')
                    if not doi and data.get('filename'):
                        parts = data['filename'].split('.')
                        if len(parts) >= 2 and '10.' in parts[1]:
                            doi = parts[1][:100]
                    
                    await asyncpg_conn.execute(
                        stmt,
                        str(uuid.uuid4()),
                        truncate_text(data.get('filename'), 255),
                        f"{category_code}/{truncate_text(data.get('filename'), 255) or 'paper.pdf'}",
                        0,
                        truncate_text(data.get('material_name'), 500),
                        json.dumps([data.get('first_author')] if data.get('first_author') else [], ensure_ascii=False),
                        truncate_text(doi, 100),
                        None,
                        json.dumps([], ensure_ascii=False),
                        None,
                        truncate_text(data.get('material_name'), 200),
                        truncate_text(data.get('symmetry_phase'), 200),
                        truncate_text(data.get('structure_descriptor'), 500),
                        truncate_text(data.get('properties'), 1000),
                        truncate_text(data.get('applications'), 500),
                        truncate_text(data.get('synthesis_method'), 500),
                        truncate_text(data.get('characterization_method'), 500),
                        truncate_text(data.get('quality_control'), 200),
                        truncate_text(data.get('first_author'), 100),
                        truncate_text(data.get('corresponding_author'), 100),
                        category_code,
                        True,
                        True,
                        str(uuid.uuid4())
                    )
                    
                    imported += 1
                    if imported % 1000 == 0:
                        print(f"  已导入: {imported}/{total}")
                        
                except Exception as e:
                    errors += 1
                    if errors <= 5:
                        print(f"第 {row_idx} 行错误: {e}")
                    continue
            
            # 更新分类计数
            await asyncpg_conn.execute(
                """UPDATE material_categories 
                   SET paper_count = (SELECT COUNT(*) FROM pdf_index WHERE category = $1) 
                   WHERE code = $1""",
                category_code
            )
        
        print(f"完成! 导入: {imported}, 错误: {errors}")
        return imported, errors
        
    except Exception as e:
        print(f"导入失败: {e}")
        return 0, 0


async def generate_mock_data(category_code, count=1000):
    """生成模拟数据用于测试"""
    print(f"\n{'='*60}")
    print(f"生成模拟数据: {MATERIAL_CATEGORIES[category_code]['name']}")
    print(f"数量: {count}")
    print('='*60)
    
    category_info = MATERIAL_CATEGORIES[category_code]
    sample_materials = category_info['sample_materials']
    
    async with engine.connect() as conn:
        raw_conn = await conn.get_raw_connection()
        asyncpg_conn = raw_conn.driver_connection
        
        # 清空该类别的现有数据
        await asyncpg_conn.execute(
            "DELETE FROM pdf_index WHERE category = $1",
            category_code
        )
        
        stmt = """
            INSERT INTO pdf_index (
                id, filename, relative_path, file_size, title, authors, doi,
                abstract, keywords, publish_year,
                material_name, symmetry_phase, structure_descriptor,
                properties, applications, synthesis_method,
                characterization_method, quality_control,
                first_author, corresponding_author,
                category, is_processed, is_indexed, file_hash,
                created_at, updated_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19, $20, $21, $22, $23, $24, NOW(), NOW())
        """
        
        imported = 0
        
        for i in range(count):
            material = sample_materials[i % len(sample_materials)]
            
            await asyncpg_conn.execute(
                stmt,
                str(uuid.uuid4()),
                f"paper_{category_code}_{i+1}.pdf",
                f"{category_code}/paper_{category_code}_{i+1}.pdf",
                1024000,
                f"Research on {material} - Sample {i+1}",
                json.dumps([f"Author_{i%100}"], ensure_ascii=False),
                f"10.1000/{category_code}.{i+1}",
                f"This is a sample abstract for {material} research.",
                json.dumps([material.lower(), "materials", "research"], ensure_ascii=False),
                2020 + (i % 5),
                material,
                "Cubic" if i % 3 == 0 else "Hexagonal" if i % 3 == 1 else "Orthorhombic",
                f"Crystal structure descriptor for {material}",
                f"High strength, good conductivity, corrosion resistant",
                f"Used in aerospace, automotive, and electronics industries",
                "Sol-gel, Hydrothermal, Sintering",
                "XRD, SEM, TEM, XPS",
                "ISO 9001 certified",
                f"Author_{i%100}",
                f"Corresponding_{i%50}",
                category_code,
                True,
                True,
                str(uuid.uuid4())
            )
            
            imported += 1
            if imported % 200 == 0:
                print(f"  已生成: {imported}/{count}")
        
        # 更新分类计数
        await asyncpg_conn.execute(
            """UPDATE material_categories 
               SET paper_count = (SELECT COUNT(*) FROM pdf_index WHERE category = $1) 
               WHERE code = $1""",
            category_code
        )
    
    print(f"完成! 生成: {imported} 条模拟数据")
    return imported


async def import_all_categories():
    """导入所有类别"""
    print("\n" + "="*60)
    print("开始导入所有材料类别数据")
    print("="*60)
    
    total_imported = 0
    total_errors = 0
    
    for category_code, info in MATERIAL_CATEGORIES.items():
        excel_file = info['excel_file']
        excel_path = os.path.join(os.path.dirname(__file__), excel_file)
        
        if os.path.exists(excel_path):
            # 从Excel导入
            imported, errors = await import_from_excel(category_code, excel_path)
            total_imported += imported
            total_errors += errors
        else:
            # 生成模拟数据
            print(f"\n文件不存在，生成模拟数据: {excel_file}")
            imported = await generate_mock_data(category_code, count=1000)
            total_imported += imported
    
    # 显示最终统计
    print("\n" + "="*60)
    print("导入完成统计")
    print("="*60)
    
    async with engine.connect() as conn:
        result = await conn.execute(text("""
            SELECT category, COUNT(*) as count 
            FROM pdf_index 
            GROUP BY category 
            ORDER BY count DESC
        """))
        rows = result.fetchall()
        
        for row in rows:
            cat_name = MATERIAL_CATEGORIES.get(row[0], {}).get('name', row[0])
            print(f"  {cat_name}: {row[1]} 条记录")
        
        result = await conn.execute(text("SELECT COUNT(*) FROM pdf_index"))
        total = result.scalar()
        print(f"\n总计: {total} 条记录")
    
    print(f"\n总导入: {total_imported}, 错误: {total_errors}")


if __name__ == "__main__":
    asyncio.run(import_all_categories())
