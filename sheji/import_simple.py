"""
简单导入Excel数据
"""

import asyncio
import uuid
import json
from datetime import datetime
from core.database import engine
from sqlalchemy import text
import openpyxl


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ['none', 'null', 'nan', '']:
            return None
        return value
    return str(value) if value else None


def truncate_text(text, max_length=500):
    if not text:
        return None
    text = str(text)
    if len(text) > max_length:
        return text[:max_length-3] + '...'
    return text


async def import_data():
    print("Loading Excel...")
    wb = openpyxl.load_workbook('合金与金属材料_二级分类.xlsx')
    ws = wb.active
    
    total = ws.max_row - 1
    print(f"Total: {total} rows")
    
    headers = [cell.value for cell in ws[1]]
    mapping = {
        'file': 'filename',
        'DOI': 'doi',
        '第一作者': 'first_author',
        '通讯作者': 'corresponding_author',
        'SPL对称相': 'symmetry_phase',
        'DSC结构描述符': 'structure_descriptor',
        'PRO属性': 'properties',
        'APL应用': 'applications',
        'SMT合成工艺方法': 'synthesis_method',
        'CMT表征方法': 'characterization_method',
        'MAT材料': 'material_name',
        '质检': 'quality_control',
    }
    
    # 获取原始连接
    async with engine.connect() as conn:
        raw_conn = await conn.get_raw_connection()
        asyncpg_conn = raw_conn.driver_connection
        
        # 清空现有数据
        await asyncpg_conn.execute("DELETE FROM pdf_index WHERE category = 'alloy_metallic'")
        print("Cleared existing data")
        
        # 准备插入语句
        stmt = """
            INSERT INTO pdf_index (
                id, filename, relative_path, file_size, title, authors, doi,
                abstract, keywords, publish_year,
                material_name, symmetry_phase, structure_descriptor,
                properties, applications, synthesis_method,
                characterization_method, quality_control,
                first_author, corresponding_author,
                category, is_processed, is_indexed, file_hash,
                created_at, updated_at
            ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12, $13, $14, $15, $16, $17, $18, $19, $20, $21, $22, $23, $24, NOW(), NOW())
        """
        
        imported = 0
        errors = 0
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            try:
                data = {mapping.get(h): clean_value(row[i]) if i < len(row) else None 
                        for i, h in enumerate(headers) if h in mapping}
                
                if not data.get('material_name'):
                    continue
                
                # 提取DOI
                doi = data.get('doi')
                if not doi and data.get('filename'):
                    parts = data['filename'].split('.')
                    if len(parts) >= 2 and '10.' in parts[1]:
                        doi = parts[1][:100]
                
                await asyncpg_conn.execute(
                    stmt,
                    str(uuid.uuid4()),
                    truncate_text(data.get('filename'), 255),
                    f"alloy_metallic/{truncate_text(data.get('filename'), 255) or 'paper.pdf'}",
                    0,
                    truncate_text(data.get('material_name'), 500),
                    json.dumps([data.get('first_author')] if data.get('first_author') else [], ensure_ascii=False),
                    truncate_text(doi, 100),
                    None,
                    json.dumps([], ensure_ascii=False),
                    None,
                    truncate_text(data.get('material_name'), 200),
                    truncate_text(data.get('symmetry_phase'), 200),
                    truncate_text(data.get('structure_descriptor'), 500),
                    truncate_text(data.get('properties'), 1000),
                    truncate_text(data.get('applications'), 500),
                    truncate_text(data.get('synthesis_method'), 500),
                    truncate_text(data.get('characterization_method'), 500),
                    truncate_text(data.get('quality_control'), 200),
                    truncate_text(data.get('first_author'), 100),
                    truncate_text(data.get('corresponding_author'), 100),
                    'alloy_metallic',
                    True,
                    True,
                    str(uuid.uuid4())
                )
                
                imported += 1
                if imported % 1000 == 0:
                    print(f"  Imported: {imported}/{total}")
                    
            except Exception as e:
                errors += 1
                if errors <= 5:
                    print(f"Error row {row_idx}: {e}")
                continue
        
        # 更新分类计数
        await asyncpg_conn.execute(
            "UPDATE material_categories SET paper_count = (SELECT COUNT(*) FROM pdf_index WHERE category = 'alloy_metallic') WHERE code = 'alloy_metallic'"
        )
    
    print(f"\nDone! Imported: {imported}, Errors: {errors}")


if __name__ == "__main__":
    asyncio.run(import_data())
