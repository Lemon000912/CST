"""
导入合金与金属材料Excel数据到PostgreSQL数据库
包含11个要素：材料名称、对称相、结构描述符、属性、应用、合成方法工艺、表征方法、质检、第一作者、DOI、通讯作者
"""

import asyncio
import uuid
import json
from datetime import datetime
from core.database import engine
from sqlalchemy import text
import openpyxl


def extract_doi(filename: str) -> str:
    """从文件名中提取DOI"""
    if not filename:
        return None
    parts = filename.split('.')
    if len(parts) >= 2:
        potential_doi = '.'.join(parts[1:])
        if '10.' in potential_doi:
            return potential_doi
    return None


def clean_value(value):
    """清理单元格值"""
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        if value.lower() in ['none', 'null', 'nan', '']:
            return None
        return value
    return str(value) if value else None


def truncate_text(text, max_length=500):
    """截断文本到指定长度"""
    if not text:
        return None
    text = str(text)
    if len(text) > max_length:
        return text[:max_length-3] + '...'
    return text


async def import_excel_data(batch_size=500):
    """导入Excel数据到数据库"""
    
    print("Loading Excel file...")
    wb = openpyxl.load_workbook('合金与金属材料_二级分类.xlsx')
    ws = wb.active
    
    total_rows = ws.max_row - 1
    print(f"Total records to import: {total_rows}")
    
    headers = [cell.value for cell in ws[1]]
    print(f"Headers: {headers}")
    
    field_mapping = {
        '一级分类': 'category',
        'file': 'filename', 
        'DOI': 'doi',
        '第一作者': 'first_author',
        '通讯作者': 'corresponding_author',
        'SPL对称相': 'symmetry_phase',
        'DSC结构描述符': 'structure_descriptor',
        'PRO属性': 'properties',
        'APL应用': 'applications',
        'SMT合成工艺方法': 'synthesis_method',
        'CMT表征方法': 'characterization_method',
        'MAT材料': 'material_name',
        '质检': 'quality_control',
        'error': 'error_field'
    }
    
    # 先统计现有数量
    async with engine.begin() as conn:
        result = await conn.execute(
            text("SELECT COUNT(*) FROM pdf_index WHERE category = 'alloy_metallic'")
        )
        existing_count = result.scalar()
        print(f"Existing alloy_metallic records: {existing_count}")
    
    imported_count = 0
    error_count = 0
    batch = []
    
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        try:
            data = {}
            for i, header in enumerate(headers):
                if header in field_mapping:
                    data[field_mapping[header]] = clean_value(row[i]) if i < len(row) else None
            
            if not data.get('material_name'):
                continue
            
            if not data.get('doi') and data.get('filename'):
                data['doi'] = extract_doi(data['filename'])
            
            record = {
                'id': str(uuid.uuid4()),
                'filename': truncate_text(data.get('filename'), 255) or f"paper_{row_idx}.pdf",
                'relative_path': f"alloy_metallic/{truncate_text(data.get('filename'), 255) or f'paper_{row_idx}.pdf'}",
                'file_size': 0,
                'title': truncate_text(data.get('material_name') or data.get('filename') or 'Unknown Title', 500),
                'authors': json.dumps([data.get('first_author')] if data.get('first_author') else [], ensure_ascii=False),
                'doi': truncate_text(data.get('doi'), 100),
                'abstract': None,
                'keywords': json.dumps([], ensure_ascii=False),
                'publish_year': None,
                'material_name': truncate_text(data.get('material_name'), 200),
                'symmetry_phase': truncate_text(data.get('symmetry_phase'), 200),
                'structure_descriptor': truncate_text(data.get('structure_descriptor'), 500),
                'properties': truncate_text(data.get('properties'), 1000),
                'applications': truncate_text(data.get('applications'), 500),
                'synthesis_method': truncate_text(data.get('synthesis_method'), 500),
                'characterization_method': truncate_text(data.get('characterization_method'), 500),
                'quality_control': truncate_text(data.get('quality_control'), 200),
                'first_author': truncate_text(data.get('first_author'), 100),
                'corresponding_author': truncate_text(data.get('corresponding_author'), 100),
                'category': 'alloy_metallic',
                'is_processed': True,
                'is_indexed': True,
                'file_hash': str(uuid.uuid4()),
            }
            
            batch.append(record)
            
            if len(batch) >= batch_size:
                success = await insert_batch(batch)
                imported_count += success
                batch = []
                if imported_count % 1000 == 0:
                    print(f"Progress: {imported_count}/{total_rows} ({imported_count/total_rows*100:.1f}%)")
            
        except Exception as e:
            error_count += 1
            if error_count <= 3:
                print(f"Error at row {row_idx}: {e}")
            continue
    
    if batch:
        success = await insert_batch(batch)
        imported_count += success
    
    # 更新分类计数
    async with engine.begin() as conn:
        await conn.execute(
            text("""
                UPDATE material_categories 
                SET paper_count = (SELECT COUNT(*) FROM pdf_index WHERE category = 'alloy_metallic')
                WHERE code = 'alloy_metallic'
            """)
        )
    
    print(f"\nImport completed!")
    print(f"  Total processed: {total_rows}")
    print(f"  Imported: {imported_count}")
    print(f"  Errors: {error_count}")


async def insert_batch(batch):
    """批量插入数据 - 每条记录使用独立事务"""
    success_count = 0
    for record in batch:
        try:
            async with engine.begin() as conn:
                await conn.execute(
                    text("""
                        INSERT INTO pdf_index (
                            id, filename, relative_path, file_size, title, authors, doi,
                            abstract, keywords, publish_year,
                            material_name, symmetry_phase, structure_descriptor,
                            properties, applications, synthesis_method,
                            characterization_method, quality_control,
                            first_author, corresponding_author,
                            category, is_processed, is_indexed, file_hash,
                            created_at, updated_at
                        ) VALUES (
                            :id, :filename, :relative_path, :file_size, :title, :authors, :doi,
                            :abstract, :keywords, :publish_year,
                            :material_name, :symmetry_phase, :structure_descriptor,
                            :properties, :applications, :synthesis_method,
                            :characterization_method, :quality_control,
                            :first_author, :corresponding_author,
                            :category, :is_processed, :is_indexed, :file_hash,
                            NOW(), NOW()
                        )
                    """),
                    record
                )
            success_count += 1
        except Exception as e:
            pass  # 静默跳过错误
    return success_count


if __name__ == "__main__":
    print("Starting Excel import...")
    print("=" * 60)
    asyncio.run(import_excel_data(batch_size=200))
